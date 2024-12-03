#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Script generates parameter files for the NASA SnowEx AK IOP March 2023
from the snow pit and depth data books.
"""

_author_ = "Megan A. Mason, NASA GSFC/SSAI"
_email_ = "megan.a.mason@nasa.gov"
_status_ = "Dvp"

# standard imports
import datetime
from datetime import timedelta
from pathlib import Path
import glob
import os
import shutil
import numpy as np
import pandas as pd
import csv
from csv import writer
import textwrap
import utm
from openpyxl import load_workbook


# local imports
# custom imports
from metadata_headers_summaryFile import metadata_headers_swe, metadata_headers_enviro
import sys
sys.path.append('/Users/mamason6/Documents/snowex/core-datasets/ground/snow-pits/run-AKIOP23/October/code/process-pits') # October dir
from parsers import generate_new_filename

# ------------------------------------------------------------------------------
# Functions

## set up ##
'''
import generate_new_filename and use to c/p and set up file structure (maybe move to more general location...ie outside of OCT22 folder?)
'''


## extract snow pit data ##

# get_metadata: pulls all metadata from top of snow pit sheet and stores in dictionary
def get_metadata(xl_file):

    d = pd.read_excel(xl_file, sheet_name='FRONT')

    # metadata
    location = d['Unnamed: 0'][1]
    site = d['Unnamed: 0'][4].split(':')[0]
    plotID = d['Unnamed: 0'][6]
    plot_num = plotID[-3:]
    date = d['Unnamed: 6'][1]
    time = d['Unnamed: 6'][4]
    zone = d['Unnamed: 17'][6]
    easting = int(d['Unnamed: 8'][6])
    northing = int(d['Unnamed: 12'][6])

    # time =

    pit_datetime=datetime.datetime.combine(date, time)
    pit_datetime_str=pit_datetime.strftime('%Y-%m-%dT%H:%M')

    # adjust SE corner coordinate to ~middle of 5x5 perimeter (ideal is -/+2.5 m but UTMs will be integers, so we'll restrict to 2 to be closest to the pit face)
    easting = easting -2
    northing = northing +2

    # convert to Lat/Lon:
    lat, lon = utm.to_latlon(easting, northing, zone, 'Northern')
    lat = round(lat, 5)
    lon = round(lon, 5)
    # lat = utm.to_latlon(easting, northing, zone, 'Northern')[0].round(5)
    # lon = utm.to_latlon(easting, northing, zone, 'Northern')[1].round(5)

    # other
    hs = d['Unnamed: 4'][6]
    observers = d['Unnamed: 8'][1]
    gps = d['Unnamed: 8'][4]
    WiseSerialNo = d['Unnamed: 6'][6]
    T_start_time = d['Unnamed: 15'][4]
    T_end_time = d['Unnamed: 17'][4]

    # vegCl = d['Unnamed: 19'][2]
    measSum = d['Unnamed: 19'][6] # could rename to co-located
    comments = d['Unnamed: 6'][27]

    # get density cutter type
    rIx = (d.iloc[:,0] == 'Density Cutter \nVolume (cc)').idxmax() #locate 'Weather:' cell in spreadsheet (row Index)
    d = d.loc[rIx:,:].reset_index(drop=True)

    cutter = d['Unnamed: 0'][3]
    swe_tube = d['Unnamed: 3'][3]

    # ~~~~Flags still need to be manually entered, but set up to catch
    if "Flag: " in str(comments):
        flags = comments.split('Flag: ')[1]
        flags = Flag.replace('\n', ' ')
    else:
        flags = None

    if site == 'CPC':
        site = 'CPCRW'

    metadata = {
        'Plot No.': plot_num,
        'Location': location,
        'Site': site,
        'PlotID': plotID, #change to pit?
        'Date': date,
        'Time': time,
        'Datetime_str': pit_datetime_str,
        'Zone': str(zone)+'N', #'N'=nortern hemisphere
        'Easting': easting,
        'Northing': northing,
        'Latitude': lat,
        'Longitude': lon,
        'HS (cm)': hs,
        'Observers': str(observers),
        'Cutter': cutter,
        'SWE Tube': swe_tube,
        'WiseSerialNo': WiseSerialNo,
        'GPS & Uncert.': gps,
        'T start Time': T_start_time,
        'T end Time': T_end_time,
        'Meas. Summary': measSum,
        'Flags': flags,
        'Pit Comments': comments
        }

    return metadata
#-------------------------------------------------------------------------------
# write_parameter_header: writes header rows to parameter files
def write_parameter_header(metadata, file_path):

    p_codes = 'n/a for this parameter'
    p_codes_strat = "Grain Size: <1mm, 1-2mm, 2-4mm, 4-6mm, >6mm; Grain Type: SH=Surface Hoar, PP=Precipitation Particles, DF=Decomposing and Fragments, RG=Rounded Grains, FC=Faceted Crystals, MF=Melt Forms, IF=Ice Formations, MFcr=Melt Crust, FCsf=Near-surface Facets, PPgp=Graupel; Hand Hardness: F=Fist, 4F=4-finger, 1F=1-finger, P=Pencil, K=Knife, I=Ice; Manual Wetness: D=Dry, M=Moist, W=Wet, V=Very Wet, S=Soaked"

    # Check if file_path contains '_stratigraphy_' and adjust p_codes header line
    if '_stratigraphy_' in str(file_path):
        p_codes = f'"{p_codes_strat}"'  # double quotes to keep commas in one column (commas below are the col 1, col 2 separator)

    with open(file_path, 'w') as f:
        f.write(f"# Location, {metadata['Location']}\n")
        f.write(f"# Site, {metadata['Site']}\n")
        f.write(f"# PitID, {metadata['PlotID']}\n") # change to pit?
        f.write(f"# Date/Local Standard Time, {metadata['Datetime_str']}\n")
        f.write(f"# UTM Zone, {metadata['Zone']}\n")
        f.write(f"# Easting, {metadata['Easting']}\n")
        f.write(f"# Northing, {metadata['Northing']}\n")
        f.write(f"# Latitude, {metadata['Latitude']}\n")
        f.write(f"# Longitude, {metadata['Longitude']}\n")
        f.write(f"# Flags, {metadata['Flags']}\n")
        f.write(f"# Pit Comments, {metadata['Pit Comments']}\n") # add the stratigraphy thing here...or later
        f.write(f"# Parameter Codes, {p_codes}\n")

#-------------------------------------------------------------------------------
# write_header_rows(): writes rows for summary SWE and Envir. file
def write_header_rows(fname_summaryFile, metadata_headers):
        with open(fname_summaryFile, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            for row in metadata_headers:
                writer.writerow(row)

#-------------------------------------------------------------------------------
# get_density() get density data from field book (if density cutter used)
def get_density(filename):
    d = pd.read_excel(filename, header=10, usecols='A:F').replace(r'^\s*$', np.nan, regex=True)
    first_double_nan = min(np.where(pd.Series(d['top\n(cm)'].isnull().values).rolling(2).sum().values == 2))[0] # generally a space at the bottom measurements
    d = d.iloc[0:first_double_nan]
    d.columns = ['# Top (cm)', '-', 'Bottom (cm)', 'Density A (kg/m3)','Density B (kg/m3)','Density C (kg/m3)']
    den_cols = ['# Top (cm)', 'Bottom (cm)', 'Density A (kg/m3)','Density B (kg/m3)','Density C (kg/m3)'] #gets rid of the '-' column

    if d['# Top (cm)'].astype(str).str.contains('Density Cutter').any():
        cutter_loc = np.argwhere(d['# Top (cm)'].str.contains('Density Cutter').values==1)
        d = d.loc[d.index[0:cutter_loc[0][0]]]
    density = d[den_cols].astype(float)
    density.dropna(axis=0, thresh=3, inplace=True)
    AvgDensity=density[['Density A (kg/m3)', 'Density B (kg/m3)', 'Density C (kg/m3)']].mean(axis=1) # not gapfilled or interpolated, but used for LWC calulation
    # print(density)
    density.to_csv(fname_density, sep=',', index=False, mode='a', na_rep=-9999) #write density csv (with NaN's)
    print('wrote: .../' + fname_density.name)
    density['Density B (kg/m3)'] = density[['Density B (kg/m3)', 'Density C (kg/m3)']].mean(axis=1) # mean of profile B with any "extra" density samples (i.e. C)
    # density.dropna(axis=0, thresh=3, inplace=True) # eventually remove, but likely bad now or we'll mesh together the 1000 and 100 cc cutters!?
    # density.dropna(axis=0, how='all', inplace=True) # eventually remove, but likely bad now or we'll mesh together the 1000 and 100 cc cutters!?
    density = density.reset_index(drop=True)
    # AvgDensity=density[['Density A (kg/m3)', 'Density B (kg/m3)', 'Density C (kg/m3)']].mean(axis=1) # not gapfilled or interpolated, but used for LWC calulation

    # print(AvgDensity)

    return density, AvgDensity

#-------------------------------------------------------------------------------
# split_dataframe: splits density df for cases where 1000 & 100 cc used
def split_dataframe(df):
    # Identify the split index
    split_index = None
    for i in range(1, len(df)):
        if df['# Top (cm)'].iloc[i] > df['# Top (cm)'].iloc[i-1]:
            split_index = i
            break  # Exits the loop, but the script continues to run

    # Split the dataframe or create 'd' if no split point is found
    if split_index is not None:
        d_1000 = df.iloc[:split_index].reset_index(drop=True)
        d_100 = df.iloc[split_index:].reset_index(drop=True)
        return d_1000, d_100
    else:
        return df, None, None
#-------------------------------------------------------------------------------
# grab_sweTube: for empty density sections, point to SWE Tube sample and grab data
def grab_sweTube(filename):

    # xl = pd.ExcelFile(filename)
    d = pd.read_excel(filename, usecols='F:J', sheet_name='FRONT')

    # get density cutter type
    rIx = (d.iloc[:,0] == 'SWE Sample').idxmax() #locate 'Weather:' cell in spreadsheet (row Index)
    d = d.iloc[rIx:,:].reset_index(drop=True)
    cols = d[['Unnamed: 7', 'Unnamed: 8', 'Unnamed: 9']]
    new_col_names = cols.iloc[0].tolist()
    cols.columns = new_col_names
    d = cols.drop([0, 1]) # drop two left columns
    first_nan = min(np.where(d['Depth \n(cm)'].isnull().values == True))[0] # get the index of the first nan value
    d = d.iloc[0:first_nan]
    d.reset_index(drop=True, inplace=True)

    # compute average to report in Summary SWE file
    swe_tube_average = d.mean()

    # Access the means using the column names
    HS = round(swe_tube_average['Depth \n(cm)']) # this will overwrite the HS at the top of the snow pit sheet.
    avgSWE = round(swe_tube_average['SWE \n(mm)'])
    avgDens = round(swe_tube_average['Density\n(kg/m3)'])

    # return HS, avgDens, avgSWE

    swe_data = {
        'Density A Mean': np.nan,
        'Density B Mean': np.nan,
        'Density Mean': avgDens,
        'SWE A': np.nan,
        'SWE B': np.nan,
        'SWE': avgSWE,
        'HS': HS # avg. SWE tube heights
    }

    return swe_data

#-------------------------------------------------------------------------------
# information to look for bottom of the pit void space (i.e. SNOW VOID written in comments)
def check_for_snow_void(filename):

    # Part I: get stratigraphy dataframe
    d = pd.read_excel(filename, sheet_name='FRONT', header=10, usecols='L:AA')
    first_double_nan = min(np.where(pd.Series(d['top\n(cm).1'].isnull().values).rolling(2).sum().values == 2))[0] # because .xlsx cells are merged
    d = d.iloc[0:first_double_nan]
    d.columns = ['# Top (cm)', '-.1', 'Bottom (cm)', 'Grain Size (mm)', '1-2 mm', '2-4 mm',
       '4-6 mm', '> 6 mm', 'Grain Type', 'Hand Hardness', 'Manual Wetness', 'Comments'] #rename them here, all of them
    strat_cols = ['# Top (cm)', 'Bottom (cm)', 'Grain Size (mm)', 'Grain Type',
                    'Hand Hardness','Manual Wetness', 'Comments'] #select which onces for the CSV file
    stratigraphy = d[strat_cols].dropna(how='all')
    # print(stratigraphy)

    # Part II: parse strat for snow voids - result is yes/no; if yes w/ void HS.
    last_row = stratigraphy.iloc[-1]
    if 'SNOW VOID:' in str(last_row['Comments']):  # Check if 'SNOW VOID:' is in the 'Comments' column of the last row
        void_result = 'Yes'
        void_HS = last_row['# Top (cm)']
        # print(void_result, void_HS)
        # return void_result, void_HS # 'yes void' and HS of void height
    else:
        void_result = 'No' # 'no void' okay to extrapolate to ground surface (0 cm)
        void_HS = None

    return void_result, void_HS

#-------------------------------------------------------------------------------
# 4. compute_swe: main function to compute SWE for Alaska March 23 data
def compute_swe(density, HeightOfSnow, void_result, filename): # where density is a dataframe
    # SWE calculation for summary file
    if not density.empty: #if 'density' isn't an empty dataframe do the following: (CHANGE TO NOTNA().all())
        SWEA_calc = [0.0] * len(density) #density.shape[0] #len(density) should also work here.
        SWEB_calc = [0.0] * len(density) #density.shape[0]
        sumSWEA=0
        sumSWEB=0
        densityA = 0
        sumDensityA = 0
        avgDensityA = 0
        densityB=0
        sumDensityB=0
        avgDensityB=0
        avgSWE = 0
        avgDens = 0

        # Steps to clean up density profiles:

        # 1. Does top of density match HS? If not, set it to HS (many are very close for AK data...)
        if density.at[0, '# Top (cm)'] != HeightOfSnow: # a few cases where top of density doesn't match HS
            density.at[0,'# Top (cm)'] = HeightOfSnow # e.g if there's a gap, extrapolate to the top.

        # 2. {Modify for Alaska, will need to consult Stratigraphy} Is the last density height 0cm? If not, set it to zero (swe is extrapolated to bottom)
        if void_result[1] is None: # if there is NO void, extrapolate to 0
            # density.at[len(density)-1, 'Bottom (cm)'] != 0
            density.at[len(density)-1, 'Bottom (cm)'] = 0 # e.g if it ends at 5cm, make it 0.
            # void_result[1] = 0 # could move this up, but I like it as 'None', definding it now for the summary file void column

        else:
            bottom_HS = min(density.at[len(density)-1, 'Bottom (cm)'], void_result[1]) # bottom of density (e.g. 5), bottom of stratigraphy (e.g 12), select 5 to extrapolate to
            density.at[len(density)-1, 'Bottom (cm)'] = bottom_HS

        # 3. Is there measurement overlap at the bottom (or elsewhere) in the snow pit? If so, shorten the bottom segment (more likely to be a worse measurement near the bottom, e.g 16-6, 12-2 --> 16-6, 6-0)
        for i in range(1, len(density)):

            if density.at[i, '# Top (cm)'] > density.at[i-1, 'Bottom (cm)']:
                density.at[i, '# Top (cm)'] = density.at[i-1, 'Bottom (cm)']

        # 4. {Might be affected by the drop na's above, but thres=3} Are there missing measurements in the duel (A, B) profile? (i.e A=235, B=NaN)
        density['Density A (kg/m3)'].fillna(density['Density B (kg/m3)'], inplace=True)
        density['Density B (kg/m3)'].fillna(density['Density A (kg/m3)'], inplace=True)

        # 5. Are there any places that need interpolation or extrapolation?
        density['Density A (kg/m3)'] = density['Density A (kg/m3)'].interpolate(method='linear', limit_direction='both')
        density['Density B (kg/m3)'] = density['Density B (kg/m3)'].interpolate(method='linear', limit_direction='both')

        # 6. Drop Density 'C' column since it's already been averaged by 'B' and will not be further used to compute SWE
        density.drop('Density C (kg/m3)', axis=1, inplace=True)
        # print(f"gapFilled Density:\n {density}")

        # 7. {not yet, need to set up filename for gapFilled} Save the density dataframe that has been gapfilled and used to compute SWE
        density.to_csv(fname_gapFilledDensity, sep=',', index=False, mode='a', na_rep=-9999)
        print('wrote: .../' + fname_gapFilledDensity.name)

        # print(f"{filename.name} gapFilled:\n {density}")

        for i in range(0, len(density)):

            densityA=density['Density A (kg/m3)'][i] #relic code assignment here, cleaner to read densityA and densityB, so it's been left as is.
            densityB=density['Density B (kg/m3)'][i]

            # Calculate SWE for each layer
            SWEA_calc[i] = (density['# Top (cm)'][i] - density['Bottom (cm)'][i])*densityA/100
            SWEB_calc[i] = (density['# Top (cm)'][i] - density['Bottom (cm)'][i])*densityB/100
            sumSWEA = round(sumSWEA + SWEA_calc[i])
            sumSWEB = round(sumSWEB + SWEB_calc[i])
            sumDensityA = sumDensityA + densityA*(density['# Top (cm)'][i] - density['Bottom (cm)'][i])
            sumDensityB = sumDensityB + densityB*(density['# Top (cm)'][i] - density['Bottom (cm)'][i])

        # calculate weighted average density
        avgDensityA = round(sumDensityA/density['# Top (cm)'][0])
        avgDensityB = round(sumDensityB/density['# Top (cm)'][0])
        avgDens = (avgDensityA + avgDensityB)/2
        avgSWE = (sumSWEA + sumSWEB)/2

        # print('Avg Dens: ', avgDens)
        # print('Avg SWE : ', avgSWE)

        # return avgDens, avgSWE
        swe_data = {
            'Density A Mean': avgDensityA,
            'Density B Mean': avgDensityB,
            'Density Mean': avgDens,
            'SWE A': sumSWEA,
            'SWE B': sumSWEB,
            'SWE': avgSWE,
            'HS': HeightOfSnow
        }

        return swe_data


#-------------------------------------------------------------------------------
# XX. get_LWC: grabs snow profile LWC and depth from pit sheet
def get_LWC(filename, fname_LWC, AvgDensity):
    # get LWC
    d = pd.read_excel(xl, header=10, usecols='A:H')#.replace(r'^\s*$', np.nan, regex=True)
    first_nan = min(np.where(d['top\n(cm)'].isnull().values == True))[0] # get the index of the first nan value
    d = d.iloc[0:first_nan]
    d.columns = ['# Top\n(cm)', '-', 'Bottom\n(cm)', 'kg/m3', 'kg/m3.1', 'kg/m3.2',
       'Permittivity A', 'Permittivity B'] # last col is temp distance, without it stuff breaks below...sloppy fix, but it's not used here.
    d = d.rename(columns={'# Top\n(cm)': '# Top (cm)', 'Bottom\n(cm)': 'Bottom (cm)'}) # rename without \n for snowex database
    lwc_cols=['# Top (cm)','Bottom (cm)','Permittivity A','Permittivity B']
    # print(d)
    LWC = d[lwc_cols].astype(float)
    LWC.insert(2, 'Avg Density (kg/m3)', AvgDensity, False)

    #Calculate LWC
    LWCA_calc = [0.0] * LWC.shape[0]
    LWCB_calc = [0.0] * LWC.shape[0]
    for i in range(0, LWC.shape[0]):
        if(pd.isna(LWC['Permittivity A'][i])):
            LWCA_calc[i] = np.nan
        if AvgDensity.isna().all():
            LWCA_calc[i] = np.nan
        else:
        # Conversion to LWC from WISe User's Manual
            wv = 0
            try: #if Density values not available for row, set LWC = NaN
                for j in range(0, 5):
                    ds = AvgDensity[i] / 1000 - wv  # Convert density to g/cm3
                    wv = (LWC['Permittivity A'][i] - 1 - (1.202 * ds) - (0.983 * ds**2)) / 21.3

                if(wv < 0):   # if computed LWC is less than zero, set it equal to zero
                    LWCA_calc[i] = 0.0
                else:
                    LWCA_calc[i] = wv * 100 #convert to percentage
            except:
                LWCA_calc[i] = np.nan

        if(pd.isna(LWC['Permittivity B'][i])):
            LWCB_calc[i] = np.nan
        if AvgDensity.isna().all():
            LWCA_calc[i] = np.nan
        else:
        # Calculate percentage LWC by volume from WISe User's Manual
            wv = 0
            try: #if Density values not available for row, set LWC = NaN
                for j in range(0, 5):
                    ds = AvgDensity[i] / 1000 - wv  # Convert density to g/cm3
                    wv = (LWC['Permittivity B'][i] - 1 - (1.202 * ds) - (0.983 * ds**2)) / 21.3

                if(wv < 0):   # if computed LWC is less than zero, set it equal to zero
                    LWCB_calc[i] = 0.0
                else:
                    LWCB_calc[i] = wv * 100 #convert to percentage
            except:
                LWCB_calc[i] = np.nan
    # Add calculated LWC values to dataframe and set number of significant digits
    LWC.insert(5, "LWC-vol A (%)", LWCA_calc , False)
    LWC.insert(6, "LWC-vol B (%)", LWCB_calc, False)
    LWC[['LWC-vol A (%)', 'LWC-vol B (%)']] = LWC[['LWC-vol A (%)', 'LWC-vol B (%)']].astype(float).round(2) # if values are floats, round them
    # AvgPerm=LWC[['Permittivity A', 'Permittivity B']].mean(axis=1)# pd.Series
    # AvgLWC=LWC[['LWC-vol A (%)', 'LWC-vol B (%)']].mean(axis=1) # pd.Series
    LWC.to_csv(fname_LWC, sep=',', index=False, mode='a', na_rep=-9999, encoding='utf-8')
    print('wrote: .../' + fname_LWC.name)

#-------------------------------------------------------------------------------
# XX. get_temperature: grabs snow profile temperature and depth from pit sheet
def get_temperature(filename, fname_temperature):
    # d = pd.read_excel(xl_file, sheet_name='FRONT')
    d = pd.read_excel(xl, header=10, usecols='I:J')#.replace(r'^\s*$', np.nan, regex=True)
    first_nan = min(np.where(d['(cm)'].isnull().values == True))[0]
    temperature = d.iloc[0:first_nan].astype(float)
    lenTemp = len(temperature.index) # looking for blank data....remove as needed
    d = pd.read_excel(xl, header=4, usecols='P:R')#.replace(r'^\s*$', np.nan, regex=True)
    last_row_value = temperature.shape[0]-1
    temperature['Time start/end'] = None # add column for start/end time (new SNEX21)
    temperature.at[0, 'Time start/end'] = d['START'][0] if not pd.isnull(d['START'][0]) else -9999
    temperature.at[last_row_value, 'Time start/end'] = d['END'][0] if not pd.isnull(d['END'][0]) else -9999
    temperature.columns = ['# Depth (cm)', 'Temperature (deg C)', 'Time start/end']
    temperature.to_csv(fname_temperature, sep=',', index=False, mode='a', na_rep=-9999)
    print('wrote: .../' + fname_temperature.name)

#-------------------------------------------------------------------------------
def get_stratigraphy(filename, fname_stratigraphy):

    d = pd.read_excel(filename, sheet_name='FRONT', header=10, usecols='L:AA')
    first_double_nan = min(np.where(pd.Series(d['top\n(cm).1'].isnull().values).rolling(2).sum().values == 2))[0] # because .xlsx cells are merged
    d = d.iloc[0:first_double_nan]
    d.columns = ['# Top (cm)', '-.1', 'Bottom (cm)', 'Grain Size (mm)', '1-2 mm', '2-4 mm',
       '4-6 mm', '> 6 mm', 'Grain Type', 'Hand Hardness', 'Manual Wetness', 'Comments'] #rename them here, all of them
    strat_cols = ['# Top (cm)', 'Bottom (cm)', 'Grain Size (mm)', 'Grain Type',
                    'Hand Hardness','Manual Wetness', 'Comments'] #select which onces for the CSV file
    if d['# Top (cm)'].astype(str).str.contains('Comments/Notes:').any():
        notes_loc = np.where(d['# Top (cm)'].str.contains('Comments/Notes:').values==1)
        d = d.loc[d.index[0:notes_loc[0][0]]]
    stratigraphy = d[strat_cols].dropna(how='all')
    stratigraphy.to_csv(fname_stratigraphy, sep=',', index=False, mode='a', na_rep=-9999, encoding='utf-8')
    # print(stratigraphy)
    print('wrote: .../' + fname_stratigraphy.name)

#-------------------------------------------------------------------------------
def get_siteDetails(filename, metadata):
    # this function also uses the metadata dictionary (location, site, plotID, etc.)

        # density sample instrument
        d = pd.read_excel(filename, sheet_name='FRONT')
        rIx = (d.iloc[:,0] == 'Density Cutter \nVolume (cc)').idxmax()
        d = d.loc[rIx:,:].reset_index(drop=True)
        cutter = d['Unnamed: 0'][3]
        swe_tube = d['Unnamed: 3'][3]

        # snow cover condition
        d = pd.read_excel(filename, usecols='A:U')
        rIx = (d.iloc[:,0] == 'Snow Cover Condition').idxmax() #locate 'Snow Cover Conditions:' cell in spreadsheet (row Index)
        d = d.loc[rIx:,:].reset_index(drop=True)
        snowCov = d['Unnamed: 4'][0].split(' ')[0] # None, Patchy, Continuous

        # weather
        d = pd.read_excel(filename, sheet_name='FRONT', usecols='O:AA')
        rIx = (d.iloc[:,0] == 'Weather Description:').idxmax() #locate 'Weather:' cell in spreadsheet (row Index)
        d = d.loc[rIx:,:].reset_index(drop=True)

        weather = str(d['Unnamed: 14'][1]).capitalize()
        precip_type = d['Unnamed: 18'][4]
        precip_rate = d['Unnamed: 18'][6]
        sky = d['Unnamed: 18'][8]
        wind = d['Unnamed: 18'][10]

        # ground cover/vegetation
        d = pd.read_excel(filename, sheet_name='FRONT', usecols='A:M')
        rIx = (d.iloc[:,0] == 'Vegetation').idxmax() #locate 'Snow Cover Conditions:' cell in spreadsheet (row Index)
        d = d.loc[rIx:,:].reset_index(drop=True)
        grdCov = d['Unnamed: 4'][1] # list of ground cover selected
        if isinstance(grdCov, float) and np.isnan(grdCov): # if the section is empty, it's a float nan. need both otherwise str wont run the isnan() arg.
            GroundVeg = np.nan
            VegHts = np.nan
            VegPrs = np.nan
        else:
            grdCov = grdCov.split(',')
            grdCov = [item.strip() for item in grdCov] # some (all?) have leading spaces
            VegType = d.iloc[3, [4,5,6,8,9,10,12]] # list of ALL cover types, not what was selected
            VegHts = d.iloc[5, [4,5,6,8,9,10,12]]
            VegPct = d.iloc[7, [4,5,6,8,9,10,12]]*100
            Veg = pd.DataFrame({'VegType': VegType.values, 'VegHts': VegHts.values, 'VegPct': VegPct.values})
            Veg['VegBool'] = Veg['VegType'].apply(lambda x: x in grdCov)
            Veg['VegBool'] = Veg['VegBool'].astype(bool) # assign as bool type
            Veg['VegHts'] = Veg['VegHts'].where((~Veg['VegBool']) | (Veg.index != 0), 0) # if 'Bare' is true, assign veg height of 0 cm
            Veg['VegHts'] = Veg['VegHts'].where(~(Veg['VegBool'] & Veg['VegHts'].isna()), -9999) # if veg type=TRUE and veg htn=NaN, assign -9999.

            # make outputs, into lists
            GroundVeg=Veg.where(Veg.VegBool).dropna().VegType.values.tolist() # list of ground veg (e.g. ['Bare', 'Shrub'])
            VegHts=Veg.where(Veg.VegBool).dropna().VegHts.values.tolist() # list of veg heights (e.g. [0, 15])
            VegPrs=Veg.where(Veg.VegBool).dropna().VegPct.values.tolist() # list of veg percents (e.g. [5, 95])

            GroundVeg = " | ".join(GroundVeg)
            VegHts = " | ".join(str(x) for x in VegHts)
            VegPrs = " | ".join(str(x) for x in VegPrs)

        d = pd.read_excel(filename, sheet_name='FRONT', usecols='A:M')
        rIx = (d.iloc[:,0] == 'Vegetation').idxmax() #locate 'Snow Cover Conditions:' cell in spreadsheet (row Index)
        d = d.loc[rIx:,:].reset_index(drop=True)

        # tussocks
        tussock_present = d['Unnamed: 4'][9]
        tuss_vert_ht = d['Unnamed: 8'][9] # if its <1, *100 (NSIDC entered as % sometimes....)
        tuss_horz_sp = d['Unnamed: 12'][9]
        # print('~~~~~TYPE:', type(tuss_vert_ht), tuss_vert_ht)
        if tuss_vert_ht is not np.nan or tuss_horz_sp is not np.nan:
            tussock_dims = "{} | {}".format(tuss_vert_ht, tuss_horz_sp)
        else:
            tussock_dims = np.nan
        # print('###TDIMS', tussock_dims)

        # tree/forest characteristics
        forest_type = d['Unnamed: 4'][11]
        deciduous_pct = d['Unnamed: 6'][13]*100
        evergreen_pct = d['Unnamed: 9'][13]*100
        tree_canopy = d['Unnamed: 4'][15]
        avg_tree_ht = d['Unnamed: 4'][17]
        if deciduous_pct is not np.nan and evergreen_pct is not np.nan:
            forest_pct = "{} | {}".format(deciduous_pct, evergreen_pct)
        else:
            forest_pct = np.nan

        # # vegetation comments section
        veg_forest_cmts = str(d['Unnamed: 0'][20])

        # ground conditions
        d = pd.read_excel(filename, sheet_name='FRONT', usecols='O:AA')
        rIx = (d.iloc[:,0] == 'Substrate').idxmax() #locate 'Substrate', but 'plot photos' to get surf. roughness
        d = d.loc[rIx:,:].reset_index(drop=True)
        grd_condition = d['Unnamed: 20'][1]
        grd_roughness = d['Unnamed: 20'][3]
        water = d['Unnamed: 20'][5]
        if water is np.nan:
            water = 'N/A'
        soil_substrate_cmts = str(d['Unnamed: 20'][8]).capitalize()
        if soil_substrate_cmts == 'Nan':
            soil_substrate_cmts = ''

        AssignedPlotComments = str(d['Unnamed: 19'][1])

      # create complete header
        index = ['# Location', '# Site', '# PitID', '# Date/Local Standard Time', '# UTM Zone', '# Easting (m)',
           '# Northing (m)', '# Latitude (deg)', '# Longitude (deg)',
            '# HS (cm)', '# Observers', '# WISe Serial No', '# GPS & Uncert.', '# Density Cutter/Instrument', '# Snow Cover Condition',
            '# Weather', '# Precip Type', '# Precip Rate', '# Sky', '# Wind',
            '# Ground Condition', '# Ground Roughness', '# Standing Water Present',
            '# Ground Vegetation/Cover', '# Vegetation Height (cm)', '# Percent Ground Cover (%)', '# Tussocks Present', '# Tussock Vert & Spacing (cm)',
            '# Forest Type', '# Percent Mixed Forest (%) (Deciduous|Evergreen)', '# Tree Canopy', '# Tree Height (m)', '# Vegetation/Forest Comments', '# Assigned Plot Cmts', '# Flags']
        column = ['value']
        df = pd.DataFrame(index=index, columns=column)
        #
        df['value'][0] = metadata.get('Location')
        df['value'][1] = metadata.get('Site')
        df['value'][2] = metadata.get('PlotID')
        df['value'][3] = metadata.get('Datetime_str')
        df['value'][4] = metadata.get('Zone')
        df['value'][5] = metadata.get('Easting')
        df['value'][6] = metadata.get('Northing')
        df['value'][7] = metadata.get('Latitude')
        df['value'][8] = metadata.get('Longitude')
        df['value'][9] = metadata.get('HS (cm)')
        df['value'][10] = metadata.get('Observers').replace('\n', ' ')
        df['value'][11] = metadata.get('WiseSerialNo')
        df['value'][12] = metadata.get('GPS & Uncert.')
        df['value'][13] = metadata.get('Cutter')
        df['value'][14] = snowCov
        df['value'][15] = weather
        df['value'][16] = precip_type
        df['value'][17] = precip_rate
        df['value'][18] = sky
        df['value'][19] = wind
        df['value'][20] = grd_condition
        df['value'][21] = grd_roughness
        df['value'][22] = water
        df['value'][23] = GroundVeg
        df['value'][24] = VegHts
        df['value'][25] = VegPrs
        df['value'][26] = tussock_present
        df['value'][27] = tussock_dims
        df['value'][28] = forest_type
        df['value'][29] = forest_pct
        df['value'][30] = tree_canopy
        df['value'][31] = avg_tree_ht
        df['value'][32] = veg_forest_cmts.replace('\n', ' ').replace('nan', '') #str(PitComments.split('Flag:')[0].replace('\n', ' ')) # removes the list of Flags if any.
        df['value'][33] = AssignedPlotComments.replace('\n', ' ').replace('nan', '')
        # df['value'][34] = Flags.replace('\n', ' ')

        df.replace('nan', np.nan, inplace=True)

        df.to_csv(fname_siteDetails, sep=',', header=False, na_rep=-9999, encoding='utf-8-sig')
        print('wrote: .../' + fname_siteDetails.name)

        # currently forgetting soil/substrate comments

        env_data = {
            'Snow Cover Condition': snowCov,
            'Precipitation Type': precip_type,
            'Precipitation Rate': precip_rate,
            'Sky': sky,
            'Wind': wind,
            'Ground Condition': grd_condition,
            'Ground Roughness': grd_roughness,
            'Standing Water Present': water,
            'Ground Vegetation': GroundVeg,
            'Height of Ground Vegetation': VegHts,
            'Percent of Ground Cover': VegPrs,
            'Tussock(s)': tussock_present,
            'Tussock Dimensions': tussock_dims,
            'Forest Type': forest_type,
            'Percent Mixed Forest': forest_pct,
            'Canopy': tree_canopy,
            'Canopy Height': avg_tree_ht
        }

        return env_data

# ------------------------------------------------------------------------------
# run main
if __name__ == "__main__":

    # static variables
    campaign_prefix = 'SnowEx23_SnowPits_AKIOP_'
    version = 'v01'

    # paths
    src_path = Path('/Users/mamason6/Documents/snowex/campaigns/AKIOP-23/march/data-management/Field_Books')
    des_basepath = Path('/Users/mamason6/Documents/snowex/core-datasets/ground/snow-pits/run-AKIOP23/March/outputs')

    summary_swe_df = []
    summary_env_df = []


    # copy raw .xlsx and place a copy in the submission package file structure
    for i, filename in enumerate(sorted(src_path.rglob('*.xlsx'))): #494WA_20230308, 684I_20230314
        print(filename.name)
        # print(filename.stem.split('_')[1])

        # get date from fname in datetime
        fdate = pd.to_datetime(filename.stem.split('_')[1], format='%Y%m%d') # stem gets rid of extention

        # reorder filename to match pitID codes (modifed by M.Mason during cleaning process, easier to sort by number!)
        first_part, second_part = filename.stem.split('_')
        new_first_part = f"{first_part[3:]}{first_part[:3]}" #494WA -->WA494, or 006CB -->CB006
        filename_stem = f"{new_first_part}_{second_part}"
        print('~~~~~~', filename_stem)

        # initialize new directories for parameter files and copy .xlsm into dir.
        new_partial_file_path, flight_line = generate_new_filename(filename) # returns path and string
        pitPath = des_basepath.joinpath('xls2csv/pits/' + new_partial_file_path)
        if not Path.exists(pitPath):
            Path(pitPath).mkdir(parents=True, exist_ok=True)

        new_filename = Path(campaign_prefix + filename_stem + '_pitSheet_' + version + filename.suffix)
        shutil.copy(filename, pitPath.joinpath(new_filename))

        # open excel pit sheet
        # xl = pitPath.joinpath(new_filename) # full filename for the copied xlsx

        # convert to standard time (-1hr if >= March 12, 2023) and resave file
        if fdate >=pd.Timestamp('2023-03-12'):
            xl = pitPath.joinpath(new_filename) # full filename for the copied xlsx
            wb = load_workbook(xl)
            ws = wb.active

            # get pit time, cell G6
            pit_time = ws['G6'].value
            print(type(pit_time))
            ws['G6'].value = pit_time.replace(hour=(pit_time.hour - 1) % 24)
            # ws['G6'].value = new_time

            # # get Temp start time (if not empty), cell P6
            temp_start_time = ws['P6'].value
            if temp_start_time is not None:
                ws['P6'].value = temp_start_time.replace(hour=(temp_start_time.hour - 1) % 24)

            # get Temp end time (if not empty), cell R6
            temp_end_time = ws['R6'].value
            if temp_end_time is not None:
                ws['R6'].value = temp_end_time.replace(hour=(temp_end_time.hour - 1) % 24)

            # # resave (or overwrite) "new_filename" .xlsx file
            wb.save(xl)

        # open excel pit sheet
        xl = pitPath.joinpath(new_filename)

        # create a dictionary of metadata from the pit sheet header
        metadata = get_metadata(xl)

        # initiate parameter file names
        fname_density         = pitPath.joinpath(campaign_prefix + filename_stem + '_density_' + version +'.csv')
        fname_gapFilledDensity= pitPath.joinpath(campaign_prefix + filename_stem + '_gapFilled_density_' + version +'.csv')
        fname_LWC             = pitPath.joinpath(campaign_prefix + filename_stem + '_LWC_' + version +'.csv')
        fname_temperature     = pitPath.joinpath(campaign_prefix + filename_stem + '_temperature_' + version +'.csv')
        fname_stratigraphy    = pitPath.joinpath(campaign_prefix + filename_stem + '_stratigraphy_' + version +'.csv')
        fname_siteDetails     = pitPath.joinpath(campaign_prefix + filename_stem + '_siteDetails_' + version +'.csv')

        # write parameter file metadata header rows
        write_parameter_header(metadata, fname_density)
        write_parameter_header(metadata, fname_gapFilledDensity)
        write_parameter_header(metadata, fname_LWC)
        write_parameter_header(metadata, fname_temperature)
        write_parameter_header(metadata, fname_stratigraphy)
        write_parameter_header(metadata, fname_siteDetails)

        # append parameter data to parameter files

        # Density and SWE
        density, AvgDensity = get_density(xl)
        HeightOfSnow = metadata.get('HS (cm)') # uses HS to adjust gapFilled_density and compute SWE if needed.
        result = split_dataframe(density)
        valid_cutters_1000cc_method = [1000, 250, '1000, 250'] # strange, but combo of int and str

        ## Handle the result ##
        #1. SWE Tube only samples
        if result[0]['Density A (kg/m3)'].isna().all(): # SWE Tube samples only, no cutters used (n=6)
            # print(result[0])
            swe_result = grab_sweTube(xl)
            void_result = check_for_snow_void(xl)


        # 2. "Regular" 1000 cc cutter sampling (majority of snow pits here)
        elif result[1] is None and metadata.get('Cutter') in valid_cutters_1000cc_method:  # No split occurred
            d_1000 = result[0]
            void_result = check_for_snow_void(xl)
            # avgDens, avgSWE = compute_swe(d_1000, void_result, xl)
            swe_result = compute_swe(d_1000, HeightOfSnow, void_result, xl)
            # print("d_1000 (or 250):")
            # print(d_1000)

        # 3. 100 cc cutter only
        elif result[1] is None:  # No split occurred and not in cutter list (i.e. 100 cc only)
            d_100 = result[0]
            void_result = check_for_snow_void(xl)
            # avgDens, avgSWE = compute_swe(d_100, void_result, xl) #(n=4)
            swe_result = compute_swe(d_100, HeightOfSnow, void_result, xl) #(n=4)
            # print("d_100 (only):")
            # print(d_100)

        # 4. Two duel profiles, 1000 & 100 cc cutter used. Only working with 1000 cc result now.
        else:  # Split occurred
            d_1000, d_100 = result
            void_result = check_for_snow_void(xl)
            # avgDens, avgSWE = compute_swe(d_1000, void_result, xl) # notice, only running for 1000 cc (add 100 cc if wanted)
            swe_result = compute_swe(d_1000, HeightOfSnow, void_result, xl) # notice, only running for 1000 cc (add 100 cc if wanted)
            # print("d_1000:")
            # print(d_1000)
            # print("\nd_100:")
            # print(d_100)


        # Liquid Water Content
        get_LWC(xl, fname_LWC, AvgDensity)

        # Temperature
        get_temperature(xl, fname_temperature) #good shape file-wise

        # Straigraphy
        get_stratigraphy(xl, fname_stratigraphy)

        # Site Details
        env_result = get_siteDetails(xl, metadata)


        summary_swe_df.append({'Plot No.': metadata.get('Plot No.'),
                            'Location': metadata.get('Location'),
                            'Site': metadata.get('Site'),
                            'PlotID': metadata.get('PlotID'),
                            'Date/Local Standard Time': metadata.get('Datetime_str'),
                            'Zone': metadata.get('Zone'),
                            'Easting': metadata.get('Easting'),
                            'Northing': metadata.get('Northing'),
                            'Latitude': metadata.get('Latitude'),
                            'Longitude': metadata.get('Longitude'),
                            'Density A Mean (kg/m^3)': swe_result['Density A Mean'],
                            'Density B Mean (kg/m^3)': swe_result['Density B Mean'],
                            'Density Mean (kg/m^3)': swe_result['Density Mean'],
                            'SWE A (mm)': swe_result['SWE A'],
                            'SWE B (mm)': swe_result['SWE B'],
                            'SWE (mm)': swe_result['SWE'],
                            'HS (cm)': swe_result['HS'], # if metadata.get('HS (cm)') used then swe tube only cases will be way off....(likely HS is greater than swe tube depth...)
                            'Snow Void (cm)': void_result[1]})


        summary_env_df.append({'Plot No.': metadata.get('Plot No.'),
                            'Location': metadata.get('Location'),
                            'Site': metadata.get('Site'),
                            'PlotID': metadata.get('PlotID'),
                            'Date/Local Standard Time': metadata.get('Datetime_str'),
                            'Zone': metadata.get('Zone'),
                            'Easting': metadata.get('Easting'),
                            'Northing': metadata.get('Northing'),
                            'Latitude': metadata.get('Latitude'),
                            'Longitude': metadata.get('Longitude'),
                            'Snow Cover Condition': env_result.get('Snow Cover Condition'),
                            'Precipitation Type': env_result.get('Precipitation Type'),
                            'Precipitation Rate': env_result.get('Precipitation Rate'),
                            'Sky': env_result.get('Sky'),
                            'Wind': env_result.get('Wind'),
                            'Ground Condition': env_result.get('Ground Condition'),
                            'Ground Roughness': env_result.get('Ground Roughness'),
                            'Standing Water Present': env_result.get('Standing Water Present'),
                            'Ground Vegetation': env_result.get('Ground Vegetation'),
                            'Height of Ground Vegetation': env_result.get('Height of Ground Vegetation'),
                            'Percent of Ground Cover': env_result.get('Percent of Ground Cover'),
                            'Tussock(s)': env_result.get('Tussock(s)'),
                            'Tussock Dimensions': env_result.get('Tussock Dimensions'),
                            'Forest Type': env_result.get('Forest Type'),
                            'Percent Mixed Forest': env_result.get('Percent Mixed Forest'),
                            'Canopy': env_result.get('Canopy'),
                            'Canopy Height': env_result.get('Canopy Height')})



    # initiate summary files
    fname_summarySWE = des_basepath.joinpath('xls2csv/'+ campaign_prefix + 'Summary_SWE_' + version + '.csv')
    fname_summaryEnviro = des_basepath.joinpath('xls2csv/'+ campaign_prefix + 'Summary_Environment_' + version + '.csv')

    # write summary file data to file
    r = write_header_rows(fname_summarySWE, metadata_headers_swe)
    r = write_header_rows(fname_summaryEnviro, metadata_headers_enviro)

    # fill in summary files
    df_SWE = pd.DataFrame(summary_swe_df)
    df_SWE.sort_values(by=['Plot No.'], inplace=True)
    df_SWE.drop('Plot No.', axis=1, inplace=True)
    df_SWE.to_csv(fname_summarySWE, mode='a', header=False, index=False) #na_rep=-9999,

    df_env = pd.DataFrame(summary_env_df)
    df_env.sort_values(by=['Plot No.'], inplace=True)
    df_env.drop('Plot No.', axis=1, inplace=True)
    df_env.to_csv(fname_summaryEnviro, mode='a', na_rep=-9999, header=False, index=False)





    print('I am done and ran smoothly')
